# -*- coding: utf-8 -*-
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

F="Arial"
BLUE=Font(name=F,size=10,color="0000FF")            # 하드코딩 입력
BLK =Font(name=F,size=10)                            # 수식
GRN =Font(name=F,size=10,color="008000")             # 타 시트 참조
H1  =Font(name=F,size=13,bold=True)
H2  =Font(name=F,size=10,bold=True,color="FFFFFF")
LBL =Font(name=F,size=10,bold=True)
HDRF=PatternFill("solid",fgColor="1F3864")
YEL =PatternFill("solid",fgColor="FFFF00")
GRY =PatternFill("solid",fgColor="F2F2F2")
thin=Side(style="thin",color="BFBFBF"); BOX=Border(thin,thin,thin,thin)
MON='#,##0;(#,##0);-'; PCT='0.0%'; MULT='0.0x'; WON='#,##0"원"'

wb=Workbook()

# ══════════════ 1. 표지·설계서 ══════════════
s=wb.active; s.title="0.설계서"
s.column_dimensions['A'].width=3
s.column_dimensions['B'].width=26
s.column_dimensions['C'].width=88
rows=[
 ("H1","테크윙(089030)·제너셈(217190) 밸류에이션 설계서",""),
 ("","기준일","2026-08-15 / 주가는 2026-08-13~14 종가"),
 ("","원천자료","DART 반기보고서(2026.06) 원문 · 테크윙 접수 20260814001211 / 제너셈 20260813000859"),
 ("","단위","억원(별도 표기 제외) · 주식수 백만주 · 주가 원"),
 ("SP","",""),
 ("H2","1. 설계 원칙",""),
 ("","P1 3계층 분리","실현(감사受 실적) / IR계획(회사 가이던스) / 공시계획(수주잔고 전량) 을 절대 섞지 않는다"),
 ("","P2 기대 분리","현재 EV − 확인가치(실현) = 제3자 기대가치. 이 값을 별도 계상해 사업단위로 환산한다"),
 ("","P3 순차입금 명시","DCF는 EV를 산출한다. 자기자본가치 = EV − 순차입금. 이 차감을 누락하지 않는다"),
 ("","P4 출처 등급","①공시원문 ②회사IR발표 ③증권사추정 ④언론추정. ③④는 밸류에이션 본체에 넣지 않는다"),
 ("","P5 반증 우선","시나리오를 지지하는 근거보다 무너뜨리는 근거를 먼저 기재한다"),
 ("SP","",""),
 ("H2","2. 밸류에이션 방법",""),
 ("","DCF","FCFF = EBIT×(1−t) + D&A − CapEx − ΔWC · 5년 명시예측 + Gordon 영구성장"),
 ("","할인율","WACC. Ke = Rf + β×ERP. 테크윙 β 1.13 / 제너셈 β 1.14 (FnGuide 52주 주간베타)"),
 ("","PER","조정순이익 기준. 테크윙은 통화선도 평가손익을 제거(2026.03 손실누계 147.4억)"),
 ("","교차검증","DCF와 PER 결과가 2배 이상 벌어지면 가정을 재검토한다"),
 ("SP","",""),
 ("H2","3. 3계층 정의",""),
 ("","① 실현","2026 상반기 확정실적 × 2. 추가 가정 없음. 가장 보수적 바닥값"),
 ("","② IR계획","회사가 공개 발표한 연간 목표. 테크윙=수주잔고 반영, 제너셈=전년비 +40%(2026.08.13 발표)"),
 ("","③ 공시계획","반기보고서 수주잔고를 연내 전량 매출인식. 테크윙 602.7억 / 제너셈 749.2억"),
 ("SP","",""),
 ("H2","4. 검증된 사실 (공시원문)",""),
 ("","계약부채(선수금)","테크윙 16.3억→322.0억(+1,875%) / 제너셈 53.5억→57.1억(+7%)"),
 ("","수주잔고","테크윙 602.7억(2026.08.07, Micron·SK하이닉스) / 제너셈 749.2억(2026.06.30)"),
 ("","순차입금","테크윙 2,719.0억 / 제너셈 168.9억"),
 ("","고객집중도","테크윙 거래처A 단독 56.3% / 제너셈 단일고객 33.9%"),
 ("","매출채권 회수일","테크윙 86일(24)→97일(25)→77일(26H) / 제너셈 48일→68일→100일"),
 ("","제너셈 대손","충당금 12.61억→9.75억(2.86억 환입), 6개월초과 장기연체 22.46억→5.13억(−77%)"),
 ("SP","",""),
 ("H2","5. 한계 (모델이 답하지 못하는 것)",""),
 ("","L1","성장률·OPM 경로는 애널리스트 가정이며 공시 근거가 없다"),
 ("","L2","테크윙 교환사채 739.9억의 교환가액·행사기간 미확인 → 희석효과 미반영"),
 ("","L3","제너셈 중국매출 76.2억의 고객 실체 미확인(SK하이닉스 충칭 추정, CXMT 아님)"),
 ("","L4","큐브프로버 대당 22.5억은 언론 추정치(20~25억)의 중간값"),
 ("","L5","본 모델은 투자자문이 아니다"),
]
r=1
for kind,b,c in rows:
    if kind=="SP": r+=1; continue
    if kind=="H1":
        s.cell(r,2,b).font=H1; r+=1; continue
    if kind=="H2":
        cc=s.cell(r,2,b); cc.font=H2; cc.fill=HDRF
        s.cell(r,3,"").fill=HDRF; r+=1; continue
    s.cell(r,2,b).font=LBL
    s.cell(r,3,c).font=BLK
    r+=1
for row in s.iter_rows(min_row=1,max_row=r,min_col=2,max_col=3):
    for cell in row: cell.alignment=Alignment(vertical="top",wrap_text=True)
print("sheet0 rows",r)

# ══════════════ 2. 가정 (입력 시트) ══════════════
a=wb.create_sheet("1.가정")
for col,w in zip("ABCDEFG",[3,30,14,14,14,14,40]): a.column_dimensions[col].width=w
def hdr(sh,r,cells):
    for i,v in enumerate(cells):
        c=sh.cell(r,2+i,v); c.font=H2; c.fill=HDRF; c.border=BOX
        c.alignment=Alignment(horizontal="center")
a.cell(1,2,"가정 입력 시트  ·  파란 글씨 = 직접 수정 / 노란 채움 = 핵심 레버").font=H1

ASSUM=[
 ("공통",None,None,None,None,None),
 ("법인세율",0.22,None,None,PCT,"한국 실효세율 가정"),
 ("영구성장률(g)",0.02,None,None,PCT,"명목GDP 하회 가정"),
 ("무위험수익률(Rf)",0.030,None,None,PCT,"국고채 기준 가정 · 확인 필요"),
 ("시장위험프리미엄(ERP)",0.060,None,None,PCT,"국내 통상치"),
 ("테크윙",None,None,None,None,None),
 ("베타(β)",1.13,None,None,'0.00',"FnGuide 52주 주간베타"),
 ("소형주 프리미엄",0.000,None,None,PCT,"KQ150 편입, 미적용"),
 ("주식수(백만주)",37.05,None,None,'0.00',"발행 37,053,645주"),
 ("순차입금",2719.0,None,None,MON,"차입 3,397.7 − 현금성 678.7 (반기보고서)"),
 ("현재주가(원)",49250,None,None,WON,"2026-08-14"),
 ("제너셈",None,None,None,None,None),
 ("베타(β)",1.14,None,None,'0.00',"FnGuide 52주 주간베타"),
 ("소형주 프리미엄",0.020,None,None,PCT,"시총 789억 · 커버리지 0"),
 ("주식수(백만주)",13.15,None,None,'0.00',"발행 13,153,761주"),
 ("순차입금",168.9,None,None,MON,"차입 209.9 − 현금 41.0 (반기보고서)"),
 ("현재주가(원)",6000,None,None,WON,"2026-08-14"),
]
r=3
for nm,v,_,_,fmt,note in ASSUM:
    if v is None:
        c=a.cell(r,2,nm); c.font=LBL; c.fill=GRY
        for k in range(3,8): a.cell(r,k,"").fill=GRY
        r+=1; continue
    a.cell(r,2,nm).font=BLK
    c=a.cell(r,3,v); c.font=BLUE; c.number_format=fmt; c.fill=YEL; c.border=BOX
    a.cell(r,7,note).font=Font(name=F,size=9,color="808080")
    r+=1
ROW={}
r=3
for nm,v,_,_,_,_ in ASSUM:
    if v is None: r+=1; continue
    ROW[nm if nm not in ROW else nm+"_2"]=r; r+=1
print(ROW)

TAX="'1.가정'!$C$4"; GT="'1.가정'!$C$5"; RF="'1.가정'!$C$6"; ERP="'1.가정'!$C$7"
CO={"테크윙":dict(beta="'1.가정'!$C$9",sp="'1.가정'!$C$10",sh="'1.가정'!$C$11",
                nd="'1.가정'!$C$12",px="'1.가정'!$C$13"),
    "제너셈":dict(beta="'1.가정'!$C$15",sp="'1.가정'!$C$16",sh="'1.가정'!$C$17",
                nd="'1.가정'!$C$18",px="'1.가정'!$C$19")}
# 3계층 시나리오 입력 (매출/OPM) — 출처 명시
TIER={"테크윙":[("① 실현","1H 1,092×2",2184,0.203),
              ("② IR계획","수주잔고 602.7 반영",2600,0.215),
              ("③ 공시계획","선수금 322 기반 상향",3100,0.225)],
      "제너셈":[("① 실현","1H 300.2×2",600,0.057),
              ("② IR계획","회사 전년비+40%",795,0.115),
              ("③ 공시계획","수주잔고 749.2 전량",1049,0.160)]}
GROW={"테크윙":[.20,.15,.10,.07,.05],"제너셈":[.25,.18,.12,.08,.05]}
OPEX={"테크윙":dict(da=.045,capex=.050,wc=.15),"제너셈":dict(da=.040,capex=.045,wc=.15)}

def build_dcf(name, sheetname):
    d=wb.create_sheet(sheetname); c=CO[name]; o=OPEX[name]
    for col,w in zip("ABCDEFGHI",[3,26,13,13,13,13,13,13,30]): d.column_dimensions[col].width=w
    d.cell(1,2,f"{name} DCF (FCFF) — 3계층").font=H1
    # WACC 블록
    d.cell(3,2,"자본비용").font=LBL; d.cell(3,2).fill=GRY
    for k in range(3,9): d.cell(3,k,"").fill=GRY
    d.cell(4,2,"Ke = Rf + β×ERP + 소형주").font=BLK
    d.cell(4,3,f"={RF}+{c['beta']}*{ERP}+{c['sp']}").font=BLK; d.cell(4,3).number_format=PCT
    d.cell(5,2,"WACC (=Ke, 무차입 가정)").font=BLK
    d.cell(5,3,"=C4").font=BLK; d.cell(5,3).number_format=PCT; d.cell(5,3).fill=YEL
    d.cell(5,9,"※ 순차입금은 EV에서 직접 차감(P3). 이중계상 방지 위해 WACC는 Ke 사용").font=Font(name=F,size=9,color="808080")
    row=7
    anchors=[]
    for label,src,rev0,opm0 in TIER[name]:
        d.cell(row,2,f"{label}  ({src})").font=H2; d.cell(row,2).fill=HDRF
        for k in range(3,9): d.cell(row,k,"").fill=HDRF
        row+=1
        hdr(d,row,["항목","기준연도","1년차","2년차","3년차","4년차","5년차"])
        row+=1
        r0=row
        d.cell(row,2,"매출").font=BLK
        cc=d.cell(row,3,rev0); cc.font=BLUE; cc.number_format=MON; cc.fill=YEL
        for i,g in enumerate(GROW[name]):
            col=4+i
            d.cell(row,col,f"={get_column_letter(col-1)}{row}*(1+{get_column_letter(col)}{row+1})").font=BLK
            d.cell(row,col).number_format=MON
        row+=1
        d.cell(row,2,"성장률").font=BLK
        for i,g in enumerate(GROW[name]):
            cc=d.cell(row,4+i,g); cc.font=BLUE; cc.number_format=PCT; cc.fill=YEL
        row+=1
        d.cell(row,2,"영업이익률").font=BLK
        cc=d.cell(row,3,opm0); cc.font=BLUE; cc.number_format=PCT; cc.fill=YEL
        for i in range(5):
            col=4+i
            cc=d.cell(row,col,opm0); cc.font=BLUE; cc.number_format=PCT; cc.fill=YEL
        row+=1
        d.cell(row,2,"영업이익(EBIT)").font=BLK
        for i in range(6):
            col=3+i
            d.cell(row,col,f"={get_column_letter(col)}{r0}*{get_column_letter(col)}{r0+2}").font=BLK
            d.cell(row,col).number_format=MON
        ebit=row; row+=1
        d.cell(row,2,"NOPAT").font=BLK
        for i in range(6):
            col=3+i
            d.cell(row,col,f"={get_column_letter(col)}{ebit}*(1-{TAX})").font=BLK
            d.cell(row,col).number_format=MON
        nopat=row; row+=1
        for lbl,ratio,sign in [("＋ 감가상각(D&A)",o['da'],1),("－ CapEx",o['capex'],-1)]:
            d.cell(row,2,lbl).font=BLK
            for i in range(6):
                col=3+i
                d.cell(row,col,f"={get_column_letter(col)}{r0}*{ratio}*{sign}").font=BLK
                d.cell(row,col).number_format=MON
            row+=1
        da_r=nopat+1; capex_r=nopat+2
        d.cell(row,2,"－ 운전자본증가").font=BLK
        d.cell(row,3,0).font=BLK; d.cell(row,3).number_format=MON
        for i in range(5):
            col=4+i
            d.cell(row,col,f"=-({get_column_letter(col)}{r0}-{get_column_letter(col-1)}{r0})*{o['wc']}").font=BLK
            d.cell(row,col).number_format=MON
        wc_r=row; row+=1
        d.cell(row,2,"FCFF").font=LBL
        for i in range(6):
            col=3+i; L=get_column_letter(col)
            d.cell(row,col,f"={L}{nopat}+{L}{da_r}+{L}{capex_r}+{L}{wc_r}").font=BLK
            d.cell(row,col).number_format=MON; d.cell(row,col).border=BOX
        fcff=row; row+=1
        d.cell(row,2,"현가계수").font=BLK
        for i in range(5):
            col=4+i
            d.cell(row,col,f"=1/(1+$C$5)^{i+1}").font=BLK; d.cell(row,col).number_format='0.000'
        df=row; row+=1
        d.cell(row,2,"FCFF 현가").font=BLK
        for i in range(5):
            col=4+i; L=get_column_letter(col)
            d.cell(row,col,f"={L}{fcff}*{L}{df}").font=BLK; d.cell(row,col).number_format=MON
        pvrow=row; row+=2
        res=row
        items=[("명시예측 PV",f"=SUM(D{pvrow}:H{pvrow})",MON),
               ("잔존가치(TV)",f"=H{fcff}*(1+{GT})/($C$5-{GT})",MON),
               ("TV 현가",f"=C{res+1}*H{df}",MON),
               ("기업가치(EV)",f"=C{res}+C{res+2}",MON),
               ("－ 순차입금",f"=-{c['nd']}",MON),
               ("자기자본가치",f"=C{res+3}+C{res+4}",MON),
               ("적정주가(원)",f"=C{res+5}/{c['sh']}*100",WON),
               ("현재가 대비",f"=C{res+6}/{c['px']}-1",PCT),
               ("TV 비중",f"=C{res+2}/C{res+3}",PCT)]
        for i,(lbl,f_,fmt) in enumerate(items):
            d.cell(res+i,2,lbl).font=LBL if i>=3 else BLK
            cc=d.cell(res+i,3,f_); cc.font=BLK; cc.number_format=fmt; cc.border=BOX
            if lbl in ("적정주가(원)","현재가 대비"): cc.fill=YEL
        anchors.append((label,res+6,res+3))
        row=res+len(items)+2
    return d,anchors

d_tw,anc_tw=build_dcf("테크윙","2.DCF_테크윙")
d_gs,anc_gs=build_dcf("제너셈","3.DCF_제너셈")
print("anchors",anc_tw,anc_gs)

# ══════════════ 4. PER ══════════════
p=wb.create_sheet("4.PER")
for col,w in zip("ABCDEFGH",[3,24,14,14,14,14,14,34]): p.column_dimensions[col].width=w
p.cell(1,2,"PER 밸류에이션 — 조정순이익 기준").font=H1
p.cell(2,2,"조정: 테크윙은 통화선도 평가손익을 영업외에서 제거하고 이자비용만 반영").font=Font(name=F,size=9,color="808080")
r=4
PERROW={}
for name,dsheet,anc,nonop,note in [
    ("테크윙","2.DCF_테크윙",anc_tw,-178,"연환산 이자비용 177.8억(반기 88.9)"),
    ("제너셈","3.DCF_제너셈",anc_gs,-6,"연환산 이자비용 5.9억(반기 2.94)")]:
    c=CO[name]
    p.cell(r,2,name).font=H2; p.cell(r,2).fill=HDRF
    for k in range(3,9): p.cell(r,k,"").fill=HDRF
    r+=1
    hdr(p,r,["계층","EBIT","영업외","조정순이익","EPS(원)","현재PER","PER25배 가치(원)"])
    r+=1
    for i,(label,_,evrow) in enumerate(anc):
        ebit_ref=f"'{dsheet}'!$C${evrow-12}"
        p.cell(r,2,label).font=BLK
        p.cell(r,3,f"={ebit_ref}").font=GRN; p.cell(r,3).number_format=MON
        cc=p.cell(r,4,nonop); cc.font=BLUE; cc.number_format=MON; cc.fill=YEL
        p.cell(r,5,f"=(C{r}+D{r})*(1-{TAX})").font=BLK; p.cell(r,5).number_format=MON
        p.cell(r,6,f"=E{r}/{c['sh']}*100").font=BLK; p.cell(r,6).number_format=WON
        p.cell(r,7,f"=IFERROR({c['px']}/F{r},0)").font=BLK; p.cell(r,7).number_format=MULT
        p.cell(r,8,f"=F{r}*25").font=BLK; p.cell(r,8).number_format=WON; p.cell(r,8).fill=YEL
        PERROW[(name,label)]=r
        r+=1
    p.cell(r,8,note).font=Font(name=F,size=9,color="808080")
    r+=2

# ══════════════ 5. 기대가치 분해 ══════════════
e=wb.create_sheet("5.기대가치분해")
for col,w in zip("ABCDEFG",[3,30,16,16,16,16,40]): e.column_dimensions[col].width=w
e.cell(1,2,"제3자 기대가치 분해  ·  현재EV − 확인가치(①실현)").font=H1
r=3
for name,dsheet,anc in [("테크윙","2.DCF_테크윙",anc_tw),("제너셈","3.DCF_제너셈",anc_gs)]:
    c=CO[name]
    e.cell(r,2,name).font=H2; e.cell(r,2).fill=HDRF
    for k in range(3,8): e.cell(r,k,"").fill=HDRF
    r+=1
    base=r
    items=[("현재 시가총액",f"={c['px']}*{c['sh']}/100",MON,"주가×주식수"),
           ("＋ 순차입금",f"={c['nd']}",MON,"반기보고서"),
           ("현재 EV",f"=C{base}+C{base+1}",MON,""),
           ("확인가치 EV (①실현)",f"='{dsheet}'!$C${anc[0][2]}",MON,"감사받은 실적만"),
           ("② IR계획 EV",f"='{dsheet}'!$C${anc[1][2]}",MON,"회사 가이던스"),
           ("③ 공시계획 EV",f"='{dsheet}'!$C${anc[2][2]}",MON,"수주잔고 전량"),
           ("제3자 기대가치",f"=C{base+2}-C{base+3}",MON,"현재EV − 확인가치"),
           ("  기대가치 비중",f"=C{base+6}/C{base+2}",PCT,""),
           ("  ③까지로 설명되는 부분",f"=C{base+5}-C{base+3}",MON,""),
           ("  ③으로도 설명 안 되는 잔여",f"=C{base+2}-C{base+5}",MON,"순수 미공시 기대"),
           ]
    for i,(lbl,f_,fmt,note) in enumerate(items):
        e.cell(r,2,lbl).font=LBL if "기대가치" in lbl or "현재 EV" in lbl else BLK
        cc=e.cell(r,3,f_); cc.font=GRN if "'" in f_ else BLK
        cc.number_format=fmt; cc.border=BOX
        if lbl=="제3자 기대가치": cc.fill=YEL
        e.cell(r,7,note).font=Font(name=F,size=9,color="808080")
        r+=1
    if name=="테크윙":
        resid=r-1
        r+=1
        e.cell(r,2,"잔여 기대의 사업단위 환산 (큐브프로버 대당 22.5억)").font=LBL; r+=1
        hdr(e,r,["적용PER","필요 순이익","필요 매출","필요 대수(대/년)"]); r+=1
        for m in (18,25,30):
            e.cell(r,2,m).font=BLUE; e.cell(r,2).number_format=MULT
            e.cell(r,3,f"=$C${resid}/B{r}").font=BLK; e.cell(r,3).number_format=MON
            e.cell(r,4,f"=C{r}/(1-{TAX})/0.225").font=BLK; e.cell(r,4).number_format=MON
            e.cell(r,5,f"=D{r}/22.5").font=BLK; e.cell(r,5).number_format='#,##0'
            r+=1
        e.cell(r,3,"※ 대당 22.5억 = 언론보도 20~25억 중간값(디일렉 2025.12)").font=Font(name=F,size=9,color="808080")
        r+=1
    r+=2
wb.save("/home/claude/val/valuation.xlsx")
print("saved")
